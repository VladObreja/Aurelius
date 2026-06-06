from __future__ import annotations

import hashlib
import json
import re
import shutil
from datetime import date, datetime
from pathlib import Path
from typing import Any

import openpyxl


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "AURELIUS_CXO_v1.30.xlsx"

DATA_DIR = ROOT / "data"
PROTOCOL_DIR = DATA_DIR / "protocol"
RESEARCH_DIR = DATA_DIR / "research"
DOCS_DIR = ROOT / "docs"
ARCHIVE_DIR = ROOT / "archive" / "original"

ID_RE = re.compile(r"\b(?:CAS|COP|CFS|HFMP|CLAB|PROD)-[A-Z0-9]+(?:-[A-Z0-9]+)?\b|\bHFMP-\d+\b")

TABLE_SHEETS = {
    "01_CPO": "cpo_entries",
    "02_CAS": "compounds",
    "03_COP": "practices",
    "04_CFS": "systems",
    "05_HFMP": "axes",
    "06_CLAB": "labs",
    "07_Daily_Tracker": "daily_tracker_template",
    "08_Lab_Tracking": "lab_tracking_template",
    "09_Events_Log": "events_log_template",
    "10_Version_History": "version_history",
    "11_Product_Table": "products",
    "12_Daily_By_Phase": "daily_schedule",
    "13_Daily_MVP_By_Phase": "daily_mvp",
}

TEXT_SHEETS = {
    "00_INTRO": "intro",
    "00_Foundation_Framework": "foundation_framework",
    "00_Project_Overview": "project_overview",
}


def ensure_dirs() -> None:
    for path in (PROTOCOL_DIR, RESEARCH_DIR, DOCS_DIR, ARCHIVE_DIR):
        path.mkdir(parents=True, exist_ok=True)


def normalize_header(value: Any) -> str:
    text = str(value or "").strip().lower()
    text = text.replace("/", " ").replace("&", " and ")
    text = re.sub(r"\([^)]*\)", "", text)
    text = re.sub(r"[^a-z0-9]+", "_", text)
    return text.strip("_") or "unnamed"


def cell_value(value: Any) -> Any:
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if value == "":
        return None
    return value


def split_csv(value: Any) -> list[str]:
    if value is None:
        return []
    if isinstance(value, (int, float)):
        return [str(value)]
    return [part.strip() for part in re.split(r",|\n", str(value)) if part.strip()]


def ids_from(value: Any) -> list[str]:
    if value is None:
        return []
    return sorted(set(ID_RE.findall(str(value))))


def as_phase_list(value: Any) -> list[str]:
    return split_csv(value)


def row_id(row: dict[str, Any], fallback_prefix: str, index: int) -> str:
    for key in ("id", "cas_id", "cop_id", "cfs_id", "axis_id", "product_id", "marker"):
        if row.get(key) is not None:
            raw = str(row[key]).strip()
            if key == "axis_id" and not raw.startswith("HFMP-"):
                return f"HFMP-{raw}"
            if key == "marker":
                return f"CLAB-{normalize_header(raw).upper()}"
            return raw
    return f"{fallback_prefix}-{index:03d}"


def table_rows(ws: Any, id_prefix: str) -> list[dict[str, Any]]:
    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        return []
    headers = [normalize_header(v) for v in rows[0]]
    output: list[dict[str, Any]] = []
    for idx, raw_row in enumerate(rows[1:], start=1):
        values = [cell_value(v) for v in raw_row]
        if not any(v is not None for v in values):
            continue
        item = {headers[i]: values[i] for i in range(min(len(headers), len(values))) if headers[i] != "unnamed"}
        item["canonical_id"] = row_id(item, id_prefix, idx)
        add_derived_fields(item)
        output.append(item)
    return output


def add_derived_fields(item: dict[str, Any]) -> None:
    for key in list(item.keys()):
        if key.endswith("_modules") or key.endswith("_axes") or key.endswith("_links") or key.endswith("_support"):
            item[f"{key}_list"] = split_csv(item.get(key))
        if key in {"phase_allowed", "phases", "phases_used", "phases_active", "phases_expressed", "phase"}:
            item[f"{key}_list"] = as_phase_list(item.get(key))
    found: set[str] = set()
    for value in item.values():
        if isinstance(value, str):
            found.update(ids_from(value))
    item["referenced_ids"] = sorted(found)


def extract_workbook() -> dict[str, Any]:
    wb = openpyxl.load_workbook(WORKBOOK, read_only=True, data_only=True)
    metadata: dict[str, Any] = {
        "source_workbook": WORKBOOK.name,
        "extracted_at": datetime.now().isoformat(timespec="seconds"),
        "sheets": [],
        "text_sections": {},
    }
    extracted: dict[str, Any] = {}

    for ws in wb.worksheets:
        metadata["sheets"].append({"name": ws.title, "rows": ws.max_row, "columns": ws.max_column})
        if ws.title in TEXT_SHEETS:
            lines = [str(row[0]).strip() for row in ws.iter_rows(values_only=True) if row and row[0] is not None]
            metadata["text_sections"][TEXT_SHEETS[ws.title]] = lines
        elif ws.title in TABLE_SHEETS:
            if ws.title in {"07_Daily_Tracker", "08_Lab_Tracking", "09_Events_Log"}:
                extracted[TABLE_SHEETS[ws.title]] = tracker_template(ws)
            else:
                extracted[TABLE_SHEETS[ws.title]] = table_rows(ws, TABLE_SHEETS[ws.title].upper())
        elif ws.title == "14_Print_View":
            extracted["print_view"] = [
                [cell_value(v) for v in row]
                for row in ws.iter_rows(values_only=True)
                if any(v is not None for v in row)
            ]
        elif ws.title == "15_RESET_TRACK":
            rows = list(ws.iter_rows(values_only=True))
            if len(rows) > 1:
                headers = [normalize_header(v) for v in rows[1]]
                extracted["reset_track"] = []
                for idx, raw_row in enumerate(rows[2:], start=1):
                    values = [cell_value(v) for v in raw_row]
                    if not any(v is not None for v in values):
                        continue
                    item = {headers[i]: values[i] for i in range(min(len(headers), len(values)))}
                    item["canonical_id"] = f"RESET-{idx:03d}"
                    add_derived_fields(item)
                    extracted["reset_track"].append(item)
        elif ws.title == "16_TEMP_PHASE_STACK":
            extracted["temp_phase_stack"] = table_rows_from_offset(ws, header_row=4, id_prefix="TEMP")

    write_json(PROTOCOL_DIR / "metadata.json", metadata)
    for name, rows in extracted.items():
        write_json(PROTOCOL_DIR / f"{name}.json", rows)
    return extracted | {"metadata": metadata}


def table_rows_from_offset(ws: Any, header_row: int, id_prefix: str) -> list[dict[str, Any]]:
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < header_row:
        return []
    headers = [normalize_header(v) for v in rows[header_row - 1]]
    output: list[dict[str, Any]] = []
    for idx, raw_row in enumerate(rows[header_row:], start=1):
        values = [cell_value(v) for v in raw_row]
        if not any(v is not None for v in values):
            continue
        item = {headers[i]: values[i] for i in range(min(len(headers), len(values))) if headers[i] != "unnamed"}
        item["canonical_id"] = row_id(item, id_prefix, idx)
        add_derived_fields(item)
        output.append(item)
    return output


def tracker_template(ws: Any) -> dict[str, Any]:
    rows = list(ws.iter_rows(values_only=True))
    headers = [str(v).strip() for v in rows[0] if v is not None] if rows else []
    return {
        "template_name": ws.title,
        "fields": [{"name": header, "key": normalize_header(header)} for header in headers],
        "rows": table_rows(ws, ws.title.upper()),
    }


def file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as fh:
        for block in iter(lambda: fh.read(1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def slugify(text: str) -> str:
    slug = normalize_header(text)
    return slug[:96].strip("_")


def research_sources() -> tuple[list[dict[str, Any]], list[dict[str, Any]]]:
    sources: list[dict[str, Any]] = []
    chunks: list[dict[str, Any]] = []
    for path in sorted(list(ROOT.glob("*.txt")) + list(ROOT.glob("*.pdf"))):
        source_id = f"SRC-{len(sources) + 1:03d}"
        title = path.stem
        source = {
            "source_id": source_id,
            "title": title,
            "filename": path.name,
            "extension": path.suffix.lower().lstrip("."),
            "bytes": path.stat().st_size,
            "sha256": file_sha256(path),
            "topics": infer_topics(title),
        }
        if path.suffix.lower() == ".txt":
            text = path.read_text(encoding="utf-8", errors="replace")
            source["character_count"] = len(text)
            source["excerpt"] = first_excerpt(text)
            chunks.extend(chunk_text(source_id, path.name, text))
        sources.append(source)
    write_json(RESEARCH_DIR / "sources.json", sources)
    write_json(RESEARCH_DIR / "chunks.json", chunks)
    return sources, chunks


def infer_topics(title: str) -> list[str]:
    patterns = {
        "photobiomodulation": r"photo|red light|light therapy|pbm",
        "pelvic_floor": r"pelvic|erectile|vascular|acoustic|tens|neuromodulation",
        "androgen_endocrine": r"androgen|steroid|testosterone|finasteride|pfs|pssd",
        "mitochondria_metabolism": r"metabolic|mitochond|nmn|bioenergetic|turnbuckle|carbon",
        "gut_oral_microbiome": r"oral|tobacco|microbiom|fungal|gut",
        "pterin_bh4": r"pterin|bh4|tetrahydrobiopterin",
        "peptides_bioregulators": r"peptide|bioregulator",
        "device_engineering": r"device|fabrication|commercial|market|diy",
    }
    found = [topic for topic, pattern in patterns.items() if re.search(pattern, title, re.I)]
    return found or ["general"]


def first_excerpt(text: str, limit: int = 500) -> str:
    cleaned = re.sub(r"\s+", " ", text).strip()
    return cleaned[:limit]


def chunk_text(source_id: str, filename: str, text: str, target_size: int = 1400) -> list[dict[str, Any]]:
    paragraphs = [p.strip() for p in re.split(r"\n\s*\n", text) if p.strip()]
    chunks: list[dict[str, Any]] = []
    buf: list[str] = []
    size = 0
    for paragraph in paragraphs:
        if len(paragraph) > target_size:
            if buf:
                chunks.append(make_chunk(source_id, filename, len(chunks) + 1, "\n\n".join(buf)))
                buf = []
                size = 0
            for part in split_long_text(paragraph, target_size):
                chunks.append(make_chunk(source_id, filename, len(chunks) + 1, part))
            continue
        if buf and size + len(paragraph) > target_size:
            chunks.append(make_chunk(source_id, filename, len(chunks) + 1, "\n\n".join(buf)))
            buf = []
            size = 0
        buf.append(paragraph)
        size += len(paragraph)
    if buf:
        chunks.append(make_chunk(source_id, filename, len(chunks) + 1, "\n\n".join(buf)))
    return chunks


def split_long_text(text: str, target_size: int) -> list[str]:
    sentences = re.split(r"(?<=[.!?])\s+", text)
    parts: list[str] = []
    buf: list[str] = []
    size = 0
    for sentence in sentences:
        if len(sentence) > target_size:
            if buf:
                parts.append(" ".join(buf).strip())
                buf = []
                size = 0
            for start in range(0, len(sentence), target_size):
                parts.append(sentence[start : start + target_size].strip())
            continue
        if buf and size + len(sentence) > target_size:
            parts.append(" ".join(buf).strip())
            buf = []
            size = 0
        buf.append(sentence)
        size += len(sentence)
    if buf:
        parts.append(" ".join(buf).strip())
    return [part for part in parts if part]


def make_chunk(source_id: str, filename: str, index: int, content: str) -> dict[str, Any]:
    chunk_id = f"{source_id}-CH-{index:03d}"
    return {
        "chunk_id": chunk_id,
        "source_id": source_id,
        "filename": filename,
        "chunk_index": index,
        "character_count": len(content),
        "referenced_protocol_ids": ids_from(content),
        "content": content,
    }


def write_json(path: Path, data: Any) -> None:
    path.write_text(json.dumps(data, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")


def archive_originals() -> None:
    for path in sorted(list(ROOT.glob("*.xlsx")) + list(ROOT.glob("*.txt")) + list(ROOT.glob("*.pdf"))):
        target = ARCHIVE_DIR / path.name
        if not target.exists():
            shutil.copy2(path, target)


def validate(extracted: dict[str, Any]) -> dict[str, Any]:
    canonical_ids: set[str] = set()
    for collection in ("compounds", "practices", "systems", "axes", "labs", "products"):
        for item in extracted.get(collection, []):
            canonical_ids.add(str(item["canonical_id"]))

    references: dict[str, list[str]] = {}
    for name, rows in extracted.items():
        if not isinstance(rows, list):
            continue
        for item in rows:
            if not isinstance(item, dict):
                continue
            source = item.get("canonical_id", name)
            for ref in item.get("referenced_ids", []):
                references.setdefault(str(source), []).append(ref)

    missing = []
    for source, refs in sorted(references.items()):
        for ref in sorted(set(refs)):
            if ref not in canonical_ids:
                missing.append({"source": source, "missing_reference": ref})

    report = {
        "canonical_id_count": len(canonical_ids),
        "reference_count": sum(len(v) for v in references.values()),
        "missing_reference_count": len(missing),
        "missing_references": missing,
    }
    write_json(DATA_DIR / "validation_report.json", report)
    return report


def md_table(rows: list[list[Any]]) -> str:
    if not rows:
        return ""
    header = rows[0]
    body = rows[1:]
    lines = [
        "| " + " | ".join(str(v or "") for v in header) + " |",
        "| " + " | ".join("---" for _ in header) + " |",
    ]
    for row in body:
        lines.append("| " + " | ".join(str(v or "").replace("\n", "<br>") for v in row) + " |")
    return "\n".join(lines)


def write_docs(extracted: dict[str, Any], sources: list[dict[str, Any]], report: dict[str, Any]) -> None:
    meta = extracted["metadata"]
    compounds = extracted.get("compounds", [])
    practices = extracted.get("practices", [])
    systems = extracted.get("systems", [])
    axes = extracted.get("axes", [])
    daily = extracted.get("daily_schedule", [])

    (DOCS_DIR / "README.md").write_text(
        "\n".join(
            [
                "# Aurelius Protocol Knowledge Base",
                "",
                "This repository stores the Aurelius protocol as structured data plus generated human-readable guides.",
                "",
                "The original workbook and research files are preserved under `archive/original/`. The canonical migrated data lives under `data/`.",
                "",
                "## Current Extraction",
                "",
                f"- Source workbook: `{meta['source_workbook']}`",
                f"- Extracted at: `{meta['extracted_at']}`",
                f"- Compounds: `{len(compounds)}`",
                f"- Practices: `{len(practices)}`",
                f"- Systems/modules: `{len(systems)}`",
                f"- HFMP axes: `{len(axes)}`",
                f"- Daily schedule entries: `{len(daily)}`",
                f"- Research sources: `{len(sources)}`",
                "",
                "## Important Note",
                "",
                "This is a research and self-tracking knowledge base, not medical advice. Protocol changes should be reviewed against safety constraints, labs, adverse events, and qualified clinical input where appropriate.",
            ]
        )
        + "\n",
        encoding="utf-8",
    )

    write_protocol_index(compounds, practices, systems, axes, report)
    write_compound_index(compounds)
    write_practice_index(practices)
    write_daily_guide(daily)
    write_research_index(sources)
    write_validation_doc(report)


def write_protocol_index(compounds: list[dict[str, Any]], practices: list[dict[str, Any]], systems: list[dict[str, Any]], axes: list[dict[str, Any]], report: dict[str, Any]) -> None:
    lines = [
        "# Protocol Index",
        "",
        "## Object Counts",
        "",
        f"- Compounds / agents: {len(compounds)}",
        f"- Practices / operations: {len(practices)}",
        f"- CFS modules: {len(systems)}",
        f"- HFMP axes: {len(axes)}",
        f"- Missing cross-references: {report['missing_reference_count']}",
        "",
        "## HFMP Axes",
        "",
    ]
    for item in axes:
        lines.append(f"- **{item.get('canonical_id')}**: {item.get('axis_name')}")
    lines.extend(["", "## CFS Modules", ""])
    for item in systems:
        lines.append(f"- **{item.get('canonical_id')}**: {item.get('system_name')}")
    (DOCS_DIR / "protocol-index.md").write_text("\n".join(lines) + "\n", encoding="utf-8")


def write_compound_index(compounds: list[dict[str, Any]]) -> None:
    rows = [["ID", "Name", "Category", "Phase Allowed", "Dose Range", "Role"]]
    for item in compounds:
        rows.append([
            item.get("canonical_id"),
            item.get("name"),
            item.get("category"),
            item.get("phase_allowed"),
            item.get("dose_range"),
            item.get("role"),
        ])
    (DOCS_DIR / "compound-index.md").write_text("# Compound Index\n\n" + md_table(rows) + "\n", encoding="utf-8")


def write_practice_index(practices: list[dict[str, Any]]) -> None:
    rows = [["ID", "Name", "Phase Allowed", "Frequency", "Timing"]]
    for item in practices:
        rows.append([
            item.get("canonical_id"),
            item.get("name"),
            item.get("phase_allowed"),
            item.get("frequency"),
            item.get("timing"),
        ])
    (DOCS_DIR / "practice-index.md").write_text("# Practice Index\n\n" + md_table(rows) + "\n", encoding="utf-8")


def write_daily_guide(daily: list[dict[str, Any]]) -> None:
    lines = ["# Daily Guide", ""]
    current_phase = None
    current_block = None
    for item in daily:
        phase = item.get("phase")
        block = item.get("time_block")
        if phase != current_phase:
            current_phase = phase
            current_block = None
            lines.extend(["", f"## {phase}", ""])
        if block != current_block:
            current_block = block
            lines.extend(["", f"### {block}", ""])
        optional = " _(optional)_" if item.get("optional") else ""
        lines.append(f"- **{item.get('what_to_do_take')}**{optional}")
        if item.get("notes"):
            lines.append(f"  - Notes: {item.get('notes')}")
        if item.get("continuation_maintenance"):
            lines.append(f"  - Continuation: {item.get('continuation_maintenance')}")
        if item.get("source_ids"):
            lines.append(f"  - Source IDs: `{item.get('source_ids')}`")
    (DOCS_DIR / "daily-guide.md").write_text("\n".join(lines).strip() + "\n", encoding="utf-8")


def write_research_index(sources: list[dict[str, Any]]) -> None:
    rows = [["ID", "Title", "Type", "Topics", "Bytes"]]
    for item in sources:
        rows.append([
            item["source_id"],
            item["title"],
            item["extension"],
            ", ".join(item["topics"]),
            item["bytes"],
        ])
    (DOCS_DIR / "research-index.md").write_text("# Research Index\n\n" + md_table(rows) + "\n", encoding="utf-8")


def write_validation_doc(report: dict[str, Any]) -> None:
    lines = [
        "# Validation Report",
        "",
        f"- Canonical IDs: {report['canonical_id_count']}",
        f"- References scanned: {report['reference_count']}",
        f"- Missing references: {report['missing_reference_count']}",
        "",
    ]
    if report["missing_references"]:
        lines.extend(["## Missing References", ""])
        for item in report["missing_references"][:200]:
            lines.append(f"- `{item['source']}` references missing `{item['missing_reference']}`")
        if len(report["missing_references"]) > 200:
            lines.append(f"- ...and {len(report['missing_references']) - 200} more.")
    (DOCS_DIR / "validation-report.md").write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> None:
    ensure_dirs()
    archive_originals()
    extracted = extract_workbook()
    sources, _chunks = research_sources()
    report = validate(extracted)
    write_docs(extracted, sources, report)
    print(
        json.dumps(
            {
                "status": "ok",
                "protocol_files": len(list(PROTOCOL_DIR.glob("*.json"))),
                "research_sources": len(sources),
                "missing_references": report["missing_reference_count"],
            },
            indent=2,
        )
    )


if __name__ == "__main__":
    main()
